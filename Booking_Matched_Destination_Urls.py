#openpyxl-A Python library to read/write Excel 2010 xlsx/xlsm files.
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
#Importing strcheckrubbish() from Check_Rubbish file.
from Check_Rubbish import strcheckrubbish

# Give the location of the xlsx file from where  we have to take input and write the mapped locations urls from Booking.com website.
path = "E:\\PyCharm Projects\\New zone maapping_Bookin.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj.cell(row=1, column=1)

# Print a particular col value
new_map_zone_list=[]
#Just initialize 2values in list.
new_map_zone_urls_list=['Url_Col','index_val']
#Find out maximum number of rows in the worksheet
max_row = sheet_obj.max_row
#Iterating values of a particular column number 3 i.e "New Mapping Zone Names" column.
for i in range(2, (max_row + 1)):
    cell_obj = sheet_obj.cell(row = i, column = 3)
    new_map_zone_list.append(cell_obj.value)

# print(len(new_map_zone_list))
# print(new_map_zone_list)

driver = webdriver.Chrome("E:/PyCharm Projects/chromedriver.exe")
cnt=0
#Iterating "New Mapping Zone Names" column values one by one.
for zone in new_map_zone_list:
    #To handle conditions like "LoulA%A9, Algarve, Portugal"
    zone = strcheckrubbish(zone)
    print(zone)
    cnt = cnt+1
    #Hitting booking.com
    driver.get('https://www.booking.com/')
    time.sleep(3)
    # Find Destination input box element
    search_box = driver.find_element_by_xpath('//input[@class="c-autocomplete__input sb-searchbox__input sb-destination__input"]')
    #print(search_box)
    search_box.clear()

    # Pass search string(destination name) on input box.
    search_box.send_keys(zone)
    time.sleep(3)
    # ##################write a response file after passing a destination name ####################
    Location_mapping_Resp=driver.page_source
    Loc_map = open("Location_mapping.html","w", encoding="utf-8")
    Loc_map.write(Location_mapping_Resp)
    # with open("Location_mapping.html", "r", encoding='utf-8') as f:
    #     text= f.read()
    ###########################################################
    #Storing suggested cities element into a list.
    suggested_dest_map = driver.find_elements_by_xpath('//ul[@aria-label="List of suggested destinations "]/li')
    #print(suggested_dest_map)
    suggest_dest=[]
    # Iterating suggested cities names element and storing its name into list.
    for dest_map in suggested_dest_map:
        suggest_dest.append(dest_map.get_attribute('data-label'))

    matched_dest_index=0
    #Matching condtition for suggested destination with input desition if matched
    for matched_dest in suggest_dest:
        up_matched_dest=strcheckrubbish(matched_dest).upper()
        up_zone=strcheckrubbish(zone).upper()
        if (up_matched_dest==up_zone) or (up_zone in up_matched_dest):
            print(up_matched_dest)
            print(up_zone)
            #Find out matched destination index number.
            print("Matched index no: {}".format(matched_dest_index))
            #Find out matched destination name <li> section
            select_li = driver.find_element_by_xpath('//li[@data-i="{}"]'.format(matched_dest_index))
            #print(select_li)
            time.sleep(5)
            #Click matched destination name <li> section
            select_li.click()
            time.sleep(5)
            break# Exit loop if destination matched.
        else:
            print("Not matched index no: {0}".format(matched_dest_index))
        matched_dest_index = matched_dest_index + 1

    ## search_button = driver.find_element_by_xpath('//button[@class="sb-searchbox__button  "]')
    search_button = driver.find_element_by_xpath('//button[@data-sb-id="main" or type="submit"]')
    search_button.click()
    time.sleep(3)
    #Getting matched destination url after selecting the destination name on Booking.com and storing it into the list.
    new_map_zone_urls_list.append(driver.current_url)
    # if cnt==1:
    #     driver.quit()
    #     break

#Writing matched destinationa urls into an Excel file.
for i in range(2, (max_row + 1)):
    cell_obj = sheet_obj.cell(row = i, column = 5)
    #new_map_zone_list.append(cell_obj.value)
    sheet_obj.cell(row=i, column=5).value = new_map_zone_urls_list[i]

wb_obj.save('E:\\PyCharm Projects\\New zone maapping_Bookin.xlsx')